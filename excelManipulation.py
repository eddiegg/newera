import openpyxl.cell,logging
from openpyxl.styles import Font, NamedStyle
logging.basicConfig(level=logging.DEBUG,format='%(levelname)s: %(message)s')
wb = openpyxl.load_workbook('./test/哈哈.xlsx')
# logging.debug(type(wb))
# logging.debug(wb.get_sheet_names())
# logging.debug(wb.get_sheet_by_name('数据维度')['A14'].value)
testSheet = wb.active
testFont = Font(size=16, bold=True, italic=True)
testStyle = NamedStyle(name='tt', font=testFont)
wb.add_named_style(testStyle)
testSheet['D2'].font = testFont
testSheet.row_dimensions[1].height = 80
# wb.create_sheet(title="测试",index=3)
# wb.remove_sheet(wb.get_sheet_by_name("啦啦"))
wb.save('./test/哈哈.xlsx')
# print(testSheet.rows)
# for cells in tuple(testSheet["A1":"C5"]):
#     for cell in cells:
#         print(cell.coordinate, cell.value)
# print(testSheet.cell(row=1,column=2).value)
# print(testSheet.max_column,testSheet.max_row)